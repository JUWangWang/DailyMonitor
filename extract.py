# ============================================================
#  extract.py  —  讀取三個 Excel，回傳結構化資料
#  經紀業務現有兩個來源：
#    1. 經紀業務當日作業_YYYYMMDD.xlsb  (Sheet: For日報)
#    2. 2 YYYYMMDD富邦證券追繳及處分金額彙總表.xls  (Sheet: Sheet1)
# ============================================================

import re
import subprocess
import tempfile
from pathlib import Path
from datetime import date
import openpyxl
import pandas as pd


# ── 工具：找指定日期的檔案 ───────────────────────────────────
def find_file(directory: Path, prefix: str, target_date: date) -> Path:
    """在 directory 找符合 prefix_YYYYMMDD.xlsx / .xlsm / .xlsb 的檔案"""
    date_str = target_date.strftime("%Y%m%d")
    patterns = [
        f"{prefix}_{date_str}.xlsx",
        f"{prefix}_{date_str}.xlsm",
        f"{prefix}_{date_str}.xlsb",
        f"{prefix}-{date_str}.xlsx",
        f"{prefix}-{date_str}.xlsm",
        f"{prefix}-{date_str}.xlsb",
    ]
    for p in patterns:
        f = directory / p
        if f.exists():
            return f

    raise FileNotFoundError(
        f"\n❌ 找不到 {target_date.strftime('%Y/%m/%d')} 的 [{prefix}] 檔案\n"
        f"   路徑：{directory}\n"
        f"   請確認檔案已放入資料夾，或改用指定日期：python main.py {target_date.strftime('%Y%m%d')}"
    )


def find_broker2_file(directory: Path, target_date: date) -> Path | None:
    """
    找追繳及處分金額彙總表。
    支援多種命名格式：
      2 20260313富邦證券追繳及處分金額彙總表.xls
      2_2026_03_13富邦證券追繳及處分金額彙總表.xls
    """
    date_str   = target_date.strftime("%Y%m%d")      # 20260313
    date_dashed = target_date.strftime("%Y.%m.%d")   # 2026.03.13

    if not directory.exists():
        return None

    for f in directory.iterdir():
        name = f.name
        if "追繳" not in name:
            continue
        if f.suffix.lower() not in (".xls", ".xlsx"):
            continue
        if date_str in name or date_dashed in name:
            return f
    return None


# ── 工具：xlsb / xls → openpyxl Workbook ────────────────────
def _convert_via_win32com(path: Path, out_dir: Path) -> Path:
    """
    用 Excel COM 把 xlsb/xls 存成 xlsx，回傳 xlsx 路徑。
    呼叫者自己處理例外。
    """
    import uuid
    import win32com.client          # pip install pywin32

    # 每次用不同檔名 → 完全避免「取代？」對話框
    xlsx_path = out_dir / f"{path.stem}_{uuid.uuid4().hex[:8]}.xlsx"

    # 萬一同名舊檔還在，先刪除
    if xlsx_path.exists():
        xlsx_path.unlink()

    xl = None
    wb = None
    try:
        xl = win32com.client.DispatchEx("Excel.Application")   # DispatchEx = 獨立 process
        xl.Visible          = False
        xl.DisplayAlerts    = False   # 關閉所有警告（含取代對話框）
        xl.AskToUpdateLinks = False
        wb = xl.Workbooks.Open(
            str(path.resolve()),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
        )
        # 51 = xlOpenXMLWorkbook (.xlsx)
        wb.SaveAs(str(xlsx_path.resolve()), FileFormat=51)
        wb.Close(False)
    finally:
        try:
            if wb  is not None: wb.Close(False)
        except Exception:
            pass
        try:
            if xl  is not None: xl.Quit()
        except Exception:
            pass

    if not xlsx_path.exists():
        raise FileNotFoundError(f"SaveAs 後找不到輸出檔：{xlsx_path}")
    return xlsx_path


def _load_any_xl(path: Path) -> openpyxl.Workbook:
    """
    讀取任意 Excel 格式並回傳 openpyxl Workbook（data_only）。

    優先順序：
      1. .xlsx → openpyxl 直接讀取
      2. .xlsb / .xls → win32com（Windows + 已安裝 Microsoft Excel）
                         pip install pywin32
      3. .xlsb / .xls → LibreOffice / soffice（Linux / macOS）
    """
    suffix = path.suffix.lower()

    # ── 1. xlsx：直接讀 ──────────────────────────────────────
    if suffix == ".xlsx":
        return openpyxl.load_workbook(str(path), data_only=True)

    out_dir = Path(tempfile.gettempdir())
    errors  = []

    # ── 2. win32com（Windows + Excel）───────────────────────
    try:
        xlsx_path = _convert_via_win32com(path, out_dir)
        return openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except ImportError:
        errors.append("win32com 未安裝（pip install pywin32）")
    except Exception as e:
        errors.append(f"win32com 失敗：{e}")
        print(f"  ⚠ win32com 轉檔失敗 ({path.name})：{e}")

    # ── 3. LibreOffice / soffice（Linux / macOS）────────────
    for lo_cmd in ["libreoffice", "soffice"]:
        try:
            subprocess.run(
                [lo_cmd, "--headless", "--convert-to", "xlsx",
                 str(path), "--outdir", str(out_dir)],
                capture_output=True, timeout=120, check=False
            )
            xlsx_path = out_dir / (path.stem + ".xlsx")
            if xlsx_path.exists():
                return openpyxl.load_workbook(str(xlsx_path), data_only=True)
        except FileNotFoundError:
            errors.append(f"{lo_cmd} 未安裝")
        except Exception as e:
            errors.append(f"{lo_cmd} 失敗：{e}")

    # ── 都失敗：列出全部原因 ─────────────────────────────────
    detail = "\n   ".join(errors)
    raise RuntimeError(
        f"\n❌ 無法開啟 {path.name}（{suffix}）\n"
        f"   嘗試過的方法：\n   {detail}\n\n"
        f"   解法（Windows）：\n"
        f"     1. 確認已安裝 Microsoft Excel\n"
        f"     2. pip install pywin32\n"
        f"   解法（Linux/macOS）：\n"
        f"     安裝 LibreOffice"
    )


# ── 工具：安全取值 ───────────────────────────────────────────
def _val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v if v is not None else 0

def _str(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

def _safe_float(v, default=0.0) -> float:
    """安全轉 float，遇 #NAME? 等錯誤回傳 default"""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("#") or s == "":
        return default
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return default

def _fmt_wan(v):
    """元 → 萬元字串，帶正負號"""
    if v is None or v == 0:
        return "0"
    wan = v / 10000
    if abs(wan) >= 10000:
        return f"{'+' if wan > 0 else ''}{wan/10000:.2f}億"
    return f"{'+' if wan > 0 else ''}{wan:,.0f}萬"

def _fmt_pct(v):
    if v is None:
        return "—"
    return f"{v*100:.1f}%"


# ════════════════════════════════════════════════════════════
#  01 + 02  市場風險（自營業務）
# ════════════════════════════════════════════════════════════
def extract_market(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Sheet1：各部室損益 ────────────────────────────────
    ws1 = wb["Sheet1"]
    data_date = str(_val(ws1, 13, 7))[:10].replace("-", "/")

    # 投資銀行處（Row 16-18）
    ib_rows = []
    for r in [16, 17]:
        ib_rows.append({
            "dept":    _str(ws1, r, 3),
            "dtd":     _val(ws1, r, 5),
            "mtd":     _val(ws1, r, 6),
            "ytd":     _val(ws1, r, 7),
            "status":  _str(ws1, r, 8),
            "m_pct":   _val(ws1, r, 15),   # col O 月損益%
            "y_pct":   _val(ws1, r, 16),   # col P 年損益%
        })
    ib_total = {
        "dept": "投資銀行處 合計",
        "dtd":  _val(ws1, 18, 5),
        "mtd":  _val(ws1, 18, 6),
        "ytd":  _val(ws1, 18, 7),
        "status": "", "m_pct": None, "y_pct": None,
    }

    # 金融交易處 策略部位（Row 21-23）
    strategy_rows = []
    for r in [21, 22]:
        strategy_rows.append({
            "dept":   _str(ws1, r, 3),
            "dtd":    _val(ws1, r, 5),
            "mtd":    _val(ws1, r, 6),
            "ytd":    _val(ws1, r, 7),
            "status": _str(ws1, r, 8),
            "m_pct":  None, "y_pct": None,
        })
    strategy_total = {
        "dept": "策略部位小計",
        "dtd":  _val(ws1, 23, 5),
        "mtd":  _val(ws1, 23, 6),
        "ytd":  _val(ws1, 23, 7),
        "status": "", "m_pct": None, "y_pct": None,
    }

    # 金融交易處 交易部位（Row 24-29）
    trade_rows = []
    for r in [24, 25, 26, 27, 28, 29]:
        dept = _str(ws1, r, 3)
        if not dept:
            continue
        trade_rows.append({
            "dept":   dept,
            "dtd":    _val(ws1, r, 5),
            "mtd":    _val(ws1, r, 6),
            "ytd":    _val(ws1, r, 7),
            "status": _str(ws1, r, 8),
            "m_pct":  _val(ws1, r, 15),   # col O
            "y_pct":  _val(ws1, r, 16),   # col P
        })
    trade_total = {
        "dept": "交易部位小計",
        "dtd":  _val(ws1, 30, 5),
        "mtd":  _val(ws1, 30, 6),
        "ytd":  _val(ws1, 30, 7),
        "status": "", "m_pct": None, "y_pct": None,
    }
    ft_total = {
        "dept": "金融交易處 合計",
        "dtd":  _val(ws1, 31, 5),
        "mtd":  _val(ws1, 31, 6),
        "ytd":  _val(ws1, 31, 7),
        "status": "", "m_pct": None, "y_pct": None,
    }

    # ── 市場風險限額使用率（Row 28-42）────────────────────
    ws_mkt = wb["市場風險限額控管表"]
    limit_rows = []
    for r in range(28, 43):
        dept = _str(ws_mkt, r, 1)
        biz  = _str(ws_mkt, r, 2)
        if not dept and not biz:
            continue
        budget_pct = _val(ws_mkt, r, 3)
        m_pct      = _val(ws_mkt, r, 4)
        y_pct      = _val(ws_mkt, r, 5)
        def _status(v):
            if v is None or v == 0:
                return "正常"
            if v >= 1.0:
                return "超限"
            if v >= 0.8:
                return "80%提醒"
            return "正常"
        limit_rows.append({
            "dept":       dept,
            "biz":        biz,
            "budget_pct": budget_pct,
            "m_pct":      m_pct,
            "y_pct":      y_pct,
            "m_status":   _status(m_pct),
            "y_status":   _status(y_pct),
        })

    # ── D3合併：單檔損失 ──────────────────────────────────
    ws_d3 = wb["D3合併"]
    d3_rows = []
    for r in range(2, 200):
        code = _str(ws_d3, r, 4)
        if not code:
            break
        loss_rate = _val(ws_d3, r, 10)
        chk       = _str(ws_d3, r, 11)
        d3_rows.append({
            "date":      str(_val(ws_d3, r, 1))[:10],
            "type":      _str(ws_d3, r, 2),
            "code":      code,
            "name":      _str(ws_d3, r, 5),
            "market":    _str(ws_d3, r, 6),
            "amount":    _val(ws_d3, r, 8),
            "pnl":       _val(ws_d3, r, 9),
            "loss_rate": loss_rate,
            "status":    chk,
            "note":      _str(ws_d3, r, 13),
        })

    d3_date = d3_rows[0]["date"] if d3_rows else data_date

    # 超限/警示計數
    m_loss_over = [r for r in limit_rows if r["m_status"] == "超限"]
    m_loss_warn = [r for r in limit_rows if r["m_status"] == "80%提醒"]
    y_loss_over = [r for r in limit_rows if r["y_status"] == "超限"]
    y_loss_warn = [r for r in limit_rows if r["y_status"] == "80%提醒"]
    loss_over   = m_loss_over + y_loss_over
    loss_warn   = m_loss_warn + y_loss_warn
    d3_over     = [r for r in d3_rows if r["status"] == "超限"]
    d3_warn     = [r for r in d3_rows if r["status"] == "80%提醒"]

    return {
        "data_date":      data_date,
        "d3_date":        d3_date,
        "ib_rows":        ib_rows,
        "ib_total":       ib_total,
        "strategy_rows":  strategy_rows,
        "strategy_total": strategy_total,
        "trade_rows":     trade_rows,
        "trade_total":    trade_total,
        "ft_total":       ft_total,
        "limit_rows":     limit_rows,
        "loss_over":      loss_over,
        "loss_warn":      loss_warn,
        "m_loss_over":    m_loss_over,
        "m_loss_warn":    m_loss_warn,
        "y_loss_over":    y_loss_over,
        "y_loss_warn":    y_loss_warn,
        "d3_rows":        d3_rows,
        "d3_over":        d3_over,
        "d3_warn":        d3_warn,
        "d3_top5":        d3_rows[:5],
    }


# ════════════════════════════════════════════════════════════
#  03  財管商品集中度
# ════════════════════════════════════════════════════════════
def extract_wm(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["總表"]

    data_date = ""
    conc = {
        "bond_inv":      {"pct": 0, "l1": 0.15, "l2": 0.20, "name": "", "rating": ""},
        "bond_noninv":   {"pct": 0, "l1": 0.05, "l2": 0.08, "name": "", "rating": ""},
        "fund":          {"pct": 0, "l1": 0.10, "l2": 0.15, "name": ""},
        "struct_target": {"pct": 0, "l1": 0.15, "l2": 0.20, "name": ""},
        "struct_target2":{"pct": 0, "l1": 0.15, "l2": 0.20, "name": ""},
        "struct_upper":  {"pct": 0, "l1": 0.20, "l2": 0.25, "name": "", "rating": ""},
        "struct_lower":  {"pct": 0, "l1": 0.10, "l2": 0.15, "name": ""},
    }

    def _row(r):
        return {
            "pct":    ws.cell(r, 7).value or 0,
            "l1":     ws.cell(r, 8).value or 0,
            "l2":     ws.cell(r, 9).value or 0,
            "alert":  str(ws.cell(r, 10).value or "").strip(),
            "name":   str(ws.cell(r, 12).value or "").strip(),
        }
    r19 = _row(19); r20 = _row(20); r23 = _row(23)
    r25 = _row(25); r26 = _row(26); r27 = _row(27)

    conc["bond_inv"].update({"pct": r19["pct"], "l1": r19["l1"] or 0.15, "l2": r19["l2"] or 0.20, "name": r19["name"]})
    conc["bond_noninv"].update({"pct": r20["pct"], "l1": r20["l1"] or 0.05, "l2": r20["l2"] or 0.08, "name": r20["name"]})
    conc["fund"].update({"pct": r23["pct"], "l1": r23["l1"] or 0.10, "l2": r23["l2"] or 0.15, "name": r23["name"]})
    conc["struct_target"].update({"pct": r25["pct"], "l1": r25["l1"] or 0.15, "l2": r25["l2"] or 0.20, "name": r25["name"]})
    conc["struct_upper"].update({"pct": r26["pct"], "l1": r26["l1"] or 0.20, "l2": r26["l2"] or 0.25, "name": r26["name"]})
    conc["struct_lower"].update({"pct": r27["pct"], "l1": r27["l1"] or 0.10, "l2": r27["l2"] or 0.15, "name": r27["name"]})

    alloc = {"bond": 0, "fund": 0, "struct": 0}
    for r in range(7, 12):
        cat = str(ws.cell(r, 13).value or "").strip()
        pct = ws.cell(r, 14).value or 0
        if "海外債" in cat:
            alloc["bond"] = pct
        elif "基金" in cat:
            alloc["fund"] = pct
        elif "結構型" in cat:
            alloc["struct"] = pct

    ws_ha = wb["高資產客戶"]
    ha = {
        "count":          ws_ha.cell(5, 6).value or 0,
        "total":          ws_ha.cell(5, 7).value or 0,
        "bb_count":       ws_ha.cell(8, 6).value or 0,
        "bb_amount":      ws_ha.cell(8, 7).value or 0,
        "offshore_count": ws_ha.cell(11, 6).value or 0,
        "offshore_amount":ws_ha.cell(11, 7).value or 0,
    }

    def _conc_status(item):
        pct = item["pct"] or 0
        l1  = item["l1"] or 0
        l2  = item["l2"] or 0
        if l2 and pct >= l2:
            return "達L2"
        if l1 and pct >= l1:
            return "達L1"
        if l1 and pct >= l1 * 0.8:
            return "接近L1"
        return "正常"

    for k in conc:
        conc[k]["status"] = _conc_status(conc[k])

    return {
        "data_date": data_date,
        "alloc":     alloc,
        "conc":      conc,
        "ha":        ha,
    }


# ════════════════════════════════════════════════════════════
#  04 + 05  經紀業務（兩個來源）
#  來源1: 經紀業務當日作業_YYYYMMDD.xlsb  Sheet: For日報
#  來源2: 2 YYYYMMDD富邦證券追繳及處分金額彙總表.xls  Sheet: Sheet1
# ════════════════════════════════════════════════════════════
def extract_broker(path: Path, path2: Path | None = None) -> dict:
    # ── 載入主檔（xlsb → xlsx → openpyxl）──────────────────
    wb = _load_any_xl(path)
    ws = wb["For日報"]

    def _v(r, c) -> float:
        """取值並安全轉 float，#NAME? / #REF! 等錯誤回傳 0"""
        return _safe_float(ws.cell(r, c).value)

    def _s(r, c) -> str:
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ""

    # ── 融資分佈（Rows 3-7：等級/比重/餘額千元/維持率小數）──
    dist_rows = []
    for r in range(3, 8):
        label = _s(r, 2)
        grade = label.replace("級", "")
        if grade not in ("A", "B", "C", "D", "E"):
            continue
        pct     = _v(r, 3)                          # 小數
        balance = _v(r, 4) * 1_000                  # 千元 → 元
        maint   = _v(r, 5) * 100                    # 小數 → %
        dist_rows.append({"grade": grade, "pct": pct, "balance": balance, "maint": maint})

    # Row 8: 總計
    total_balance = _v(8, 4) * 1_000               # 千元 → 元
    total_maint   = _v(8, 5) * 100                  # 小數 → %
    abc_pct       = _v(9, 3)                        # 小數

    # ── 不限用途借款分佈（Rows 14-18）──────────────────────
    unlim_dist_rows = []
    for r in range(14, 19):
        label = _s(r, 2)
        grade = label.replace("級", "")
        if grade not in ("A", "B", "C", "D", "E"):
            continue
        pct     = _v(r, 3)                          # 小數
        balance = _v(r, 4)                          # 元
        maint   = _v(r, 5) * 100                    # 小數 → %
        unlim_dist_rows.append({"grade": grade, "pct": pct, "balance": balance, "maint": maint})

    # Row 19: 不限用途總計
    unlim_total_balance = _v(12, 4) * 10_000        # 萬元 → 元（D12：不限用途借款總餘額）
    unlim_total_maint   = _v(19, 5) * 100           # 小數 → %
    unlim_abc_pct       = _v(20, 3)                 # 小數

    # ── 款項借貸餘額（Rows 25-27，單位：元）────────────────
    loans = {
        "half_year": _v(25, 3),
        "t5":        _v(26, 3),
        "t30":       _v(27, 3),
    }

    # ── 有價證券借貸（Rows 31-35，單位：元）────────────────
    sec_lending = {
        "foreign": _v(31, 3),    # 外資
        "broker":  _v(32, 3),    # 他家券商
        "prop":    _v(33, 3),    # 自營
        "nature":  _v(34, 3),    # 自然人
        "total":   _v(35, 3),    # 總計
        "rate":    0,             # 費率不對外揭露
    }

    # ── 融資前5大個股（Rows 3-7，Cols H=8 I=9 J=10 K=11 L=12）
    margin_top5 = []
    for r in range(3, 8):
        code  = _s(r, 8)
        name  = _s(r, 9)
        if not code or not name:
            continue
        grade   = _s(r, 10)
        balance = _v(r, 11) * 1e8    # 億 → 元
        maint   = _v(r, 12) * 100    # 小數 → %
        margin_top5.append({"code": code, "name": name, "grade": grade,
                            "balance": balance, "maint": maint})

    # ── 融券前5大個股（Rows 12-16，Cols H=8 I=9 J=10 K=11 L=12）
    short_top5 = []
    for r in range(12, 17):
        code  = _s(r, 8)
        name  = _s(r, 9)
        if not code or not name:
            continue
        grade  = _s(r, 10)
        collat = _v(r, 11) * 1e8    # 億 → 元
        maint  = _v(r, 12) * 100    # 小數 → %
        short_top5.append({"code": code, "name": name, "grade": grade,
                           "collat": collat, "maint": maint})

    # ── 不限用途借款前5大客戶（Rows 21-25，Cols H=8 I=9 J=10 K=11）
    unlim_top5 = []
    for r in range(21, 26):
        branch = _s(r, 8)
        name   = _s(r, 9)
        if not name:
            continue
        amount = _v(r, 10) * 10_000  # 萬 → 元
        maint  = _v(r, 11) * 100     # 小數 → %
        unlim_top5.append({"branch": branch, "name": name,
                           "amount": amount, "maint": maint})

    # ── 公司客戶違約（Row 39：C=件數, D=差額損失金額元）──────
    default_count = int(_v(39, 3) or 0)
    default_loss  = _v(39, 4)          # 元

    # ── 來源2：追繳及處分金額彙總表 ─────────────────────────
    margin_call = {
        "recall_count":   0,
        "recall_amount":  0,    # 元
        "dispose_count":  0,
        "dispose_amount": 0,    # 元
    }
    if path2 and path2.exists():
        try:
            wb2 = _load_any_xl(path2)
            ws2 = wb2["Sheet1"]
            margin_call["recall_amount"]  = _safe_float(ws2.cell(8,  2).value) * 1_000   # 千元 → 元
            margin_call["recall_count"]   = int(_safe_float(ws2.cell(9,  2).value))
            margin_call["dispose_amount"] = _safe_float(ws2.cell(10, 2).value) * 1_000   # 千元 → 元
            margin_call["dispose_count"]  = int(_safe_float(ws2.cell(11, 2).value))
        except Exception as e:
            print(f"  ⚠ 追繳處分檔讀取失敗：{e}")

    return {
        # 融資分佈
        "dist_rows":            dist_rows,
        "total_balance":        total_balance,
        "total_maint":          total_maint,
        "abc_pct":              abc_pct,
        # 不限用途借款分佈
        "unlim_dist_rows":      unlim_dist_rows,
        "unlim_total_balance":  unlim_total_balance,
        "unlim_total_maint":    unlim_total_maint,
        "unlim_abc_pct":        unlim_abc_pct,
        # 款項借貸
        "loans":                loans,
        # 有價證券借貸
        "sec_lending":          sec_lending,
        # 前5大
        "margin_top5":          margin_top5,
        "short_top5":           short_top5,
        "unlim_top5":           unlim_top5,
        # 違約
        "default_count":        default_count,
        "default_loss":         default_loss,
        # 追繳處分（來源2）
        "margin_call":          margin_call,
    }


# ════════════════════════════════════════════════════════════
#  彙整：一次讀所有 Excel
# ════════════════════════════════════════════════════════════
def extract_all(target_date: date, config) -> dict:
    from config import (BROKER_DIR, BROKER2_DIR,
                        BROKER_PREFIX,
                        MARKET_DIR, MARKET_PREFIX,
                        WM_DIR, WM_PREFIX)

    broker_file  = find_file(BROKER_DIR, BROKER_PREFIX, target_date)
    broker2_file = find_broker2_file(BROKER2_DIR, target_date)
    market_file  = find_file(MARKET_DIR, MARKET_PREFIX, target_date)
    wm_file      = find_file(WM_DIR,     WM_PREFIX,     target_date)

    print(f"  [經紀主檔]  {broker_file.name}")
    print(f"  [追繳處分]  {broker2_file.name if broker2_file else '（未找到，跳過）'}")
    print(f"  [市場]      {market_file.name}")
    print(f"  [財管]      {wm_file.name}")

    market = extract_market(market_file)
    wm     = extract_wm(wm_file)
    broker = extract_broker(broker_file, broker2_file)

    return {"market": market, "wm": wm, "broker": broker,
            "report_date": target_date.strftime("%Y/%m/%d")}
